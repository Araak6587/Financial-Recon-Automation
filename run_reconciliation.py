"""
FILE: run_reconciliation.py
PURPOSE: Run all SQL reconciliation checks and export results to Excel
This is the "automation" piece - what you'd schedule to run every EOD
"""

import sqlite3
import pandas as pd
from datetime import datetime
import os

DB_PATH = "recon.db"
OUTPUT_PATH = "output/daily_recon_report.xlsx"

os.makedirs("output", exist_ok=True)

def load_csvs(conn):
    tables = {
        "securities_master": "data/securities_master.csv",
        "bloomberg_prices":  "data/bloomberg_prices.csv",
        "internal_prices":   "data/internal_prices.csv",
        "transactions":      "data/transactions.csv",
        "gl_entries":        "data/gl_entries.csv",
        "budget_vs_actuals": "data/budget_vs_actuals.csv",
    }
    for table, path in tables.items():
        df = pd.read_csv(path)
        df.to_sql(table, conn, if_exists="replace", index=False)
    print("✓ All tables loaded")

def run_pricing_recon(conn):
    sql = """
    WITH price_comparison AS (
        SELECT
            b.price_date,
            b.security_id,
            sm.security_name,
            sm.asset_class,
            b.close_price                                        AS bloomberg_price,
            i.close_price                                        AS internal_price,
            CASE WHEN i.close_price IS NULL THEN NULL
                 ELSE ROUND((i.close_price - b.close_price)
                      / b.close_price * 100, 4)
            END                                                  AS pct_diff
        FROM bloomberg_prices b
        LEFT JOIN internal_prices i
            ON b.price_date = i.price_date AND b.security_id = i.security_id
        LEFT JOIN securities_master sm ON b.security_id = sm.security_id
    )
    SELECT *,
        CASE
            WHEN internal_price IS NULL             THEN 'MISSING PRICE'
            WHEN pct_diff > 10                      THEN 'CRITICAL'
            WHEN pct_diff > 5                       THEN 'HIGH'
            WHEN pct_diff > 1                       THEN 'MEDIUM'
            ELSE                                         'OK'
        END AS severity
    FROM price_comparison
    ORDER BY
        CASE WHEN internal_price IS NULL THEN 0
             WHEN pct_diff > 10 THEN 1
             WHEN pct_diff > 5  THEN 2
             WHEN pct_diff > 1  THEN 3
             ELSE 4 END,
        price_date
    """
    return pd.read_sql(sql, conn)

def run_daily_pricing_summary(conn):
    sql = """
    WITH cmp AS (
        SELECT
            b.price_date,
            CASE WHEN i.close_price IS NULL THEN 1 ELSE 0 END AS missing,
            CASE WHEN i.close_price IS NOT NULL
                  AND ABS(i.close_price - b.close_price)/b.close_price*100 > 1
                 THEN 1 ELSE 0 END AS discrepancy,
            CASE WHEN i.close_price IS NOT NULL
                  AND ABS(i.close_price - b.close_price)/b.close_price*100 > 10
                 THEN 1 ELSE 0 END AS critical
        FROM bloomberg_prices b
        LEFT JOIN internal_prices i
            ON b.price_date = i.price_date AND b.security_id = i.security_id
    )
    SELECT
        price_date,
        COUNT(*)            AS total_securities,
        SUM(missing)        AS missing_prices,
        SUM(discrepancy)    AS discrepancies,
        SUM(critical)       AS critical_issues,
        ROUND((SUM(missing)+SUM(discrepancy))*100.0/COUNT(*),2) AS error_rate_pct
    FROM cmp
    GROUP BY price_date
    ORDER BY price_date
    """
    return pd.read_sql(sql, conn)

def run_gl_recon(conn):
    sql = """
    WITH t AS (
        SELECT txn_date AS dt, account_id, ROUND(SUM(amount),2) AS txn_flow
        FROM transactions GROUP BY txn_date, account_id
    ),
    g AS (
        SELECT entry_date AS dt, account_id, ROUND(SUM(debit)-SUM(credit),2) AS gl_flow
        FROM gl_entries GROUP BY entry_date, account_id
    )
    SELECT
        COALESCE(t.dt, g.dt)            AS recon_date,
        COALESCE(t.account_id, g.account_id) AS account_id,
        t.txn_flow,
        g.gl_flow,
        ROUND(COALESCE(t.txn_flow,0) - COALESCE(g.gl_flow,0), 2) AS break_amount,
        CASE
            WHEN t.txn_flow IS NULL     THEN 'NO TRANSACTIONS'
            WHEN g.gl_flow IS NULL      THEN 'NO GL ENTRY'
            WHEN ABS(COALESCE(t.txn_flow,0)-COALESCE(g.gl_flow,0)) < 0.01 THEN 'RECONCILED'
            WHEN ABS(COALESCE(t.txn_flow,0)-COALESCE(g.gl_flow,0)) > 10000 THEN 'CRITICAL BREAK'
            ELSE 'BREAK'
        END AS status
    FROM t
    FULL OUTER JOIN g ON t.dt=g.dt AND t.account_id=g.account_id
    ORDER BY ABS(COALESCE(t.txn_flow,0)-COALESCE(g.gl_flow,0)) DESC
    """
    return pd.read_sql(sql, conn)

def run_budget_variance(conn):
    sql = """
    SELECT
        month, department, category,
        ROUND(budget_amount,2)                              AS budget,
        ROUND(actual_amount,2)                              AS actual,
        ROUND(actual_amount - budget_amount, 2)             AS variance,
        ROUND((actual_amount-budget_amount)/budget_amount*100,1) AS variance_pct,
        CASE WHEN actual_amount <= budget_amount THEN 'Favorable' ELSE 'Unfavorable' END AS type,
        CASE
            WHEN ABS((actual_amount-budget_amount)/budget_amount*100) > 15 THEN 'MATERIAL'
            WHEN ABS((actual_amount-budget_amount)/budget_amount*100) > 5  THEN 'NOTABLE'
            ELSE 'OK'
        END AS materiality
    FROM budget_vs_actuals
    ORDER BY ABS(actual_amount-budget_amount) DESC
    """
    return pd.read_sql(sql, conn)

def style_sheet(writer, df, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    wb = writer.book
    ws = writer.sheets[sheet_name]

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font   = Font(name="Arial", size=10)
    border_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(bottom=border_side)

    severity_colors = {
        "CRITICAL": "FF0000", "MISSING PRICE": "FF6600",
        "HIGH": "FF9900", "MEDIUM": "FFFF00",
        "CRITICAL BREAK": "FF0000", "BREAK": "FF9900",
        "MATERIAL": "FF6600", "UNFAVORABLE": "FFE0E0",
    }

    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(col)) + 4)

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.border = thin_border
            val = str(cell.value).upper() if cell.value else ""
            for keyword, color in severity_colors.items():
                if keyword in val:
                    cell.fill = PatternFill("solid", fgColor=color)
                    if color in ["FF0000", "FF6600", "1F4E79"]:
                        cell.font = Font(name="Arial", size=10, color="FFFFFF")
                    break

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def main():
    conn = sqlite3.connect(DB_PATH)
    print("Loading data...")
    load_csvs(conn)

    print("Running reconciliations...")
    pricing_detail  = run_pricing_recon(conn)
    pricing_summary = run_daily_pricing_summary(conn)
    gl_recon        = run_gl_recon(conn)
    budget_var      = run_budget_variance(conn)

    print(f"Writing report to {OUTPUT_PATH}...")
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        style_sheet(writer, pricing_summary, "Pricing Summary")
        style_sheet(writer, pricing_detail,  "Pricing Detail")
        style_sheet(writer, gl_recon,        "GL Reconciliation")
        style_sheet(writer, budget_var,      "Budget Variance")

        # Summary dashboard sheet
        summary_data = {
            "Check": ["Pricing – Total Issues", "Pricing – Critical", "GL Breaks", "Budget Material Variances"],
            "Count": [
                int((pricing_detail["severity"] != "OK").sum()),
                int((pricing_detail["severity"] == "CRITICAL").sum()),
                int((gl_recon["status"].isin(["BREAK","CRITICAL BREAK"])).sum()),
                int((budget_var["materiality"] == "MATERIAL").sum()),
            ],
            "Status": ["", "", "", ""]
        }
        import pandas as pd2
        summary_df = pd.DataFrame(summary_data)
        summary_df["Status"] = summary_df["Count"].apply(
            lambda x: "✓ CLEAN" if x == 0 else "⚠ NEEDS REVIEW"
        )
        style_sheet(writer, summary_df, "Dashboard")

    conn.close()
    print(f"\n✓ Report generated: {OUTPUT_PATH}")
    print(f"  Pricing issues:   {int((pricing_detail['severity'] != 'OK').sum())}")
    print(f"  GL breaks:        {int((gl_recon['status'].isin(['BREAK','CRITICAL BREAK'])).sum())}")
    print(f"  Budget material:  {int((budget_var['materiality'] == 'MATERIAL').sum())}")

if __name__ == "__main__":
    main()
